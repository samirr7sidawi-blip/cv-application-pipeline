"""
One-shot migration: copy every row from the OneDrive Excel tracker into the
Google Sheet so the two trackers start in sync.

Run this ONCE on your Mac (where OneDrive is mounted) after you've set up
Google service-account credentials. After that, the pipeline writes to BOTH
trackers in parallel and they stay in sync.

Usage:
    python3 tools/migrate_excel_to_sheets.py            # safe: shows what would happen
    python3 tools/migrate_excel_to_sheets.py --apply    # actually wipes + writes

Requires:
    GOOGLE_CREDENTIALS_FILE in .env (and the file must exist on disk)
    OneDrive synced locally with 'job applications tracker.xlsx'

What it does:
    1. Reads every non-empty row from the Excel tracker
    2. Opens (or creates) the Google Sheet from GOOGLE_SHEETS_ID
    3. CLEARS the existing Google Sheet content (the old format is incompatible)
    4. Writes Excel headers + every Excel row into the Sheet
"""

import argparse
import os
import sys
from datetime import datetime, date
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

# Reuse the sheet plumbing from log_to_sheets.py to avoid duplicating logic
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from log_to_sheets import (  # noqa: E402
    HEADERS, get_client, get_or_create_sheet,
    gmail_search_url, hyperlink_formula, apply_compact_formatting,
)

ONEDRIVE_DOCS = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Documents"
)
TRACKER_PATH = os.path.join(ONEDRIVE_DOCS, "job applications tracker.xlsx")


def cell_to_string(value) -> str:
    """Excel cells can be datetime/int/float/None/str — normalize for Sheets."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def inline_cover_letter_if_file_url(value: str) -> str:
    """The OneDrive Excel stores cover letters as file:// URLs pointing to .txt
    files. Google Sheets doesn't make file:// links clickable in browsers, so
    inline the actual text content if we can find the file on disk."""
    from urllib.parse import unquote, urlparse
    if not value or not value.startswith("file://"):
        return value
    try:
        path = unquote(urlparse(value).path)
        if not os.path.exists(path):
            return value
        with open(path, "r", encoding="utf-8") as f:
            text = f.read().strip()
        return text or value
    except Exception:
        return value


def read_excel_rows() -> tuple[list[str], list[list[str]]]:
    if not os.path.exists(TRACKER_PATH):
        raise SystemExit(f"Excel tracker not found at {TRACKER_PATH}")

    wb = load_workbook(TRACKER_PATH, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    found_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
    found_headers = [cell_to_string(h) for h in found_headers]

    if found_headers != HEADERS:
        print(f"WARNING: Excel headers don't match expected format.")
        print(f"  Expected: {HEADERS}")
        print(f"  Found:    {found_headers}")
        print(f"  Migration will still copy 9 columns positionally.")

    cover_letter_idx = HEADERS.index("Cover Letter")
    cv_version_idx = HEADERS.index("CV Version")
    company_idx = HEADERS.index("Company")
    position_idx = HEADERS.index("Position")

    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)[: len(HEADERS)]
        while len(cells) < len(HEADERS):
            cells.append(None)
        if any(c is not None and str(c).strip() for c in cells):
            stringified = [cell_to_string(c) for c in cells]
            # Inline the actual cover letter text (replaces unclickable file:// URLs)
            stringified[cover_letter_idx] = inline_cover_letter_if_file_url(stringified[cover_letter_idx])
            # Replace the CV Version file:// URL with a Gmail-search link that opens
            # the email containing both PDF + cover letter attachments
            email_url = gmail_search_url(stringified[position_idx], stringified[company_idx])
            if email_url:
                stringified[cv_version_idx] = hyperlink_formula(email_url, "📧 Open email")
            rows.append(stringified)

    return found_headers, rows


def main() -> int:
    parser = argparse.ArgumentParser(description="Migrate Excel tracker → Google Sheets")
    parser.add_argument("--apply", action="store_true", help="Actually write to Google Sheets (default is dry-run)")
    args = parser.parse_args()

    creds_path = os.getenv("GOOGLE_CREDENTIALS_FILE", "credentials.json")
    if not os.path.exists(creds_path):
        print(f"ERROR: Google service-account credentials not found at '{creds_path}'.")
        print("Follow the setup steps in workflows/cloud_apply.md or deploy/README.md before running this.")
        return 1

    _excel_headers, excel_rows = read_excel_rows()
    print(f"Read {len(excel_rows)} data row(s) from Excel tracker:")
    for i, r in enumerate(excel_rows, start=1):
        company = r[0] or "(blank)"
        position = r[1] or "(blank)"
        date_applied = r[2] or "(blank)"
        print(f"  {i:2}. {date_applied}  {company} — {position}")

    if not args.apply:
        print("\nDRY RUN — no changes made. Re-run with --apply to write to Google Sheets.")
        return 0

    print("\nConnecting to Google Sheets...")
    gc = get_client()
    ws = get_or_create_sheet(gc)

    print(f"Clearing existing content in worksheet '{ws.title}'...")
    ws.clear()

    payload = [HEADERS] + excel_rows
    print(f"Writing {len(payload)} rows (1 header + {len(excel_rows)} data)...")
    ws.update(values=payload, range_name="A1", value_input_option="USER_ENTERED")

    print("Applying compact formatting to long-text columns...")
    apply_compact_formatting(ws)

    sheet_id = os.getenv("GOOGLE_SHEETS_ID", "")
    print("\nMigration complete.")
    if sheet_id:
        print(f"  Sheet URL: https://docs.google.com/spreadsheets/d/{sheet_id}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
