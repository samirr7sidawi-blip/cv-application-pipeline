"""
Append one row to the user's OneDrive 'job applications tracker.xlsx'.
Also archives the tailored CV PDF and cover letter alongside the tracker
on OneDrive, and links them from the row via file:// URLs.

Usage:
    python3 tools/log_to_excel.py

Requires:
    .tmp/job_description.json
    .tmp/tailored_cv.pdf
    .tmp/cover_letter_final.txt

Output:
    Appends one row to the OneDrive tracker.
    Copies CV + cover letter into OneDrive archive folders.
"""

import json
import os
import re
import shutil
from datetime import datetime, date
from openpyxl import load_workbook
from urllib.parse import quote

ONEDRIVE_DOCS = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Documents"
)
TRACKER_PATH = os.path.join(ONEDRIVE_DOCS, "job applications tracker.xlsx")
CV_ARCHIVE = os.path.join(ONEDRIVE_DOCS, "cv_archive")
COVER_ARCHIVE = os.path.join(ONEDRIVE_DOCS, "cover_letters_archive")

EXPECTED_HEADERS = [
    "Company", "Position", "Date Applied", "Status",
    "CV Version", "Job Description", "Job Link",
    "Interview Notes", "Cover Letter",
]

STATUS_APPLIED = "🟡 applied"


def load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_text(path: str, default: str = "") -> str:
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def sanitize_filename(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9_\- ]", "", s).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:80] or "Unknown"


def file_url(path: str) -> str:
    return "file://" + quote(path, safe="/:")


def find_next_empty_row(ws) -> int:
    """First row (>= 2) where column A (Company) is empty."""
    for row_idx in range(2, ws.max_row + 2):
        val = ws.cell(row=row_idx, column=1).value
        if val is None or (isinstance(val, str) and not val.strip()):
            return row_idx
    return ws.max_row + 1


def archive_files(job: dict) -> tuple[str, str]:
    os.makedirs(CV_ARCHIVE, exist_ok=True)
    os.makedirs(COVER_ARCHIVE, exist_ok=True)

    today = date.today().strftime("%Y-%m-%d")
    base = f"{today}_{sanitize_filename(job.get('company', 'unknown'))}_{sanitize_filename(job.get('title', 'role'))[:40]}"

    cv_dest = os.path.join(CV_ARCHIVE, f"{base}.pdf")
    cover_dest = os.path.join(COVER_ARCHIVE, f"{base}.txt")

    if os.path.exists(".tmp/tailored_cv.pdf"):
        shutil.copy2(".tmp/tailored_cv.pdf", cv_dest)
    else:
        cv_dest = ""

    if os.path.exists(".tmp/cover_letter_final.txt"):
        shutil.copy2(".tmp/cover_letter_final.txt", cover_dest)
    else:
        cover_dest = ""

    return cv_dest, cover_dest


def excel_has_file_open(path: str) -> bool:
    """True if Microsoft Excel currently holds an open handle to the tracker.
    When Excel has the file open, openpyxl writes silently lose to OneDrive
    sync (Excel's in-memory copy wins). Detected via `lsof`."""
    import subprocess
    try:
        out = subprocess.run(
            ["lsof", path], capture_output=True, text=True, timeout=5
        )
        return "Microsoft" in out.stdout or "Excel" in out.stdout
    except Exception:
        return False


def main():
    if not os.path.exists(TRACKER_PATH):
        raise SystemExit(
            f"Tracker not found at {TRACKER_PATH}\n"
            f"Make sure OneDrive is synced and the file exists."
        )

    if excel_has_file_open(TRACKER_PATH):
        raise SystemExit(
            f"Microsoft Excel currently has the tracker open:\n"
            f"  {TRACKER_PATH}\n\n"
            f"Writes will be silently overwritten by OneDrive sync. "
            f"Please close the file in Excel (or quit Excel) and re-run."
        )

    job = load_json(".tmp/job_description.json")

    cv_path, cover_path = archive_files(job)
    print(f"Archived CV:           {cv_path}")
    print(f"Archived cover letter: {cover_path}")

    wb = load_workbook(TRACKER_PATH)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    headers = [ws.cell(row=1, column=i + 1).value for i in range(len(EXPECTED_HEADERS))]
    if headers != EXPECTED_HEADERS:
        print(f"Warning: column headers don't exactly match expected.")
        print(f"  Expected: {EXPECTED_HEADERS}")
        print(f"  Found:    {headers}")
        print("Proceeding anyway with positional writes (A-I).")

    row_idx = find_next_empty_row(ws)
    print(f"Writing to row {row_idx}")

    today_dt = datetime.combine(date.today(), datetime.min.time())

    ws.cell(row=row_idx, column=1, value=job.get("company", ""))
    ws.cell(row=row_idx, column=2, value=job.get("title", ""))
    ws.cell(row=row_idx, column=3, value=today_dt)
    ws.cell(row=row_idx, column=4, value=STATUS_APPLIED)
    ws.cell(row=row_idx, column=5, value=file_url(cv_path) if cv_path else "")
    ws.cell(row=row_idx, column=6, value=(job.get("description", "") or "")[:32000])
    ws.cell(row=row_idx, column=7, value=job.get("url", "") or job.get("apply_url", ""))
    ws.cell(row=row_idx, column=8, value="")
    ws.cell(row=row_idx, column=9, value=file_url(cover_path) if cover_path else "")

    wb.save(TRACKER_PATH)

    print(f"\nApplication logged to OneDrive tracker:")
    print(f"  Path: {TRACKER_PATH}")
    print(f"  Row {row_idx}: {job.get('title')} at {job.get('company')}")


if __name__ == "__main__":
    main()
